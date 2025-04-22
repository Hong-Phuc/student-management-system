import openpyxl
import sqlite3
from tkinter import filedialog, messagebox
import re

def extract_subject_name(full_subject_name):
    cleaned_name = re.sub(r"\[.*?\]|\(.*?\)|-", "", full_subject_name).strip()
    return cleaned_name

def import_excel(app):
    """Nhập dữ liệu từ file Excel"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", ".xlsx")])
    if not file_path:
        return

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        semester = sheet['C6'].value
        raw_subject_name = sheet['C9'].value
        subject_name = extract_subject_name(raw_subject_name)
        class_code = sheet['C10'].value

        student_data = []
        absence_dates = []  # Danh sách chứa các thông tin ngày vắng

        # Lấy thông tin ngày vắng từ dòng 12
        for col in range(7, 24, 3):  # GHI, JKL, MNO, PQR, STU, VWX (cột 7-23)
            date_info = sheet.cell(row=12, column=col).value
            if date_info:
                absence_dates.append(date_info.split(' - ')[-1])  # Lấy phần ngày từ chuỗi

        for row in sheet.iter_rows(min_row=14, values_only=True):
            if row[0] is not None:
                student_name = f"{row[2]} {row[3]}"
                mssv = row[1]
                phan_tram_vang = row[27]

                # Duyệt qua các cột ST
                thoi_gian_nghi = []
                for i, value in enumerate(row[7:24:3]):  # ST (H, K, N, Q, T, W)
                    if value == 5:  # 5 tiết vắng
                        date = absence_dates[i] if i < len(absence_dates) else ""
                        if date:
                            thoi_gian_nghi.append(f"{date} - vắng 5 tiết")

                so_ngay_nghi = len(thoi_gian_nghi)  # Số ngày vắng
                thoi_gian_nghi_str = " ; ".join(thoi_gian_nghi) if thoi_gian_nghi else ""

                student_data.append((semester, class_code, subject_name, student_name, mssv, so_ngay_nghi, phan_tram_vang, thoi_gian_nghi_str))

        with sqlite3.connect('data.db') as conn:
            c = conn.cursor()
            for student in student_data:
                c.execute(
                    """INSERT OR REPLACE INTO students 
                    (semester, class_code, subject_name, student_name, mssv, so_ngay_nghi, phan_tram_vang, thoi_gian_nghi) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    student
                )

        app.load_students()  # Cập nhật lại danh sách sinh viên
        
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở file Excel: {str(e)}")
